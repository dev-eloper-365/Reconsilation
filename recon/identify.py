"""Work out what an uploaded file actually is, so the upload slots don't have
to be right.

Every export in this project announces itself: the letterhead names the company
whose book it is, and the header row's own labels say which export template it
came from. Reading both is more reliable than trusting a filename, an extension,
or which box on a form the file was dropped into.
"""
import openpyxl
import xlrd

MAGIC = {"xlsx": b"PK\x03\x04", "xls": b"\xd0\xcf\x11\xe0"}

KIND_LABELS = {
    "delta_ledger": "Delta ledger (Ledger Account export)",
    "jindal_ledger": "Jindal ledger (Ledger Statement export)",
    "register": "Delta columnar register (with Quantity / Value columns)",
}


def file_format(blob):
    for kind, magic in MAGIC.items():
        if blob.startswith(magic):
            return kind
    return None


def _grid_xlsx(path, rows=40, cols=30):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=rows, max_col=cols, values_only=True)):
        grid.append(["" if v is None else str(v).strip() for v in row])
    return grid


def _grid_xls(path, rows=40, cols=30):
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    return [
        [str(sh.cell_value(r, c)).strip() for c in range(min(cols, sh.ncols))]
        for r in range(min(rows, sh.nrows))
    ]


def identify(path, blob=None):
    """-> {kind, letterhead, fmt, header_labels}. kind is None when nothing matches."""
    fmt = file_format(blob) if blob is not None else ("xlsx" if path.endswith("x") else "xls")
    try:
        grid = _grid_xlsx(path) if fmt == "xlsx" else _grid_xls(path)
    except Exception as e:
        return {"kind": None, "letterhead": "", "fmt": fmt, "header_labels": [], "error": str(e)}

    letterhead = grid[0][0] if grid and grid[0] else ""

    header, kind = [], None
    for row in grid:
        if not row:
            continue
        labels = set(row)
        if row[0] == "Voucher No.":
            header, kind = row, "jindal_ledger"
            break
        # A register's header need not start in column A: a sheet merging several
        # years prefixes "Financial Year" / "Source File". Anchor on the payload
        # columns, and only then fall back to the plain ledger's shape.
        if {"Value", "Gross Total"} <= labels:
            header, kind = row, "register"
            break
        if row[0] == "Date":
            header = row
            if "Debit" in labels and "Credit" in labels:
                kind = "delta_ledger"
            break

    # A merged sheet has its header on row 1, so "row 1 column A" is a column
    # name, not a company. Don't report a header label as a letterhead.
    if letterhead and header and letterhead in {str(h).strip() for h in header if h}:
        letterhead = ""

    return {"kind": kind, "letterhead": letterhead, "fmt": fmt,
            "header_labels": [h for h in header if h]}
