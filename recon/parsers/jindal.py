"""Parser for Jindal's Tally "Ledger Statement" export.

One workbook can hold a sheet per counterparty — a group's several companies
each get their own ledger with Jindal. Those are different legal entities and
must not be merged, so `parse` reads one sheet: the one whose name best matches
the counterparty asked for, or the first ledger sheet when none is named.
`sheets()` lists what a workbook actually contains.

Handles both the legacy `.xls` (BIFF, via xlrd) and `.xlsx` (via openpyxl)
flavours — the same export template arrives in either format depending on who
saved it, and the only real difference is how dates come back: an xlrd serial
float versus a datetime. See profiles/jindal.md for the column layout.
"""
import datetime

import openpyxl
import xlrd

MAGIC_XLSX = b"PK\x03\x04"


def _is_xlsx(path):
    with open(path, "rb") as f:
        return f.read(4) == MAGIC_XLSX


def _names(path):
    if _is_xlsx(path):
        return openpyxl.load_workbook(path, data_only=True, read_only=True).sheetnames
    return xlrd.open_workbook(path, on_demand=True).sheet_names()


def _header_row(grid):
    for i, row in enumerate(grid):
        if str(_cell(row, 0) or "").strip() == "Voucher No.":
            return i
    return None


def sheets(path):
    """-> [{name, rows}] for every sheet that looks like a ledger statement."""
    out = []
    for name in _names(path):
        grid = _grid_xlsx(path, name) if _is_xlsx(path) else _grid_xls(path, name)
        header = _header_row(grid)
        if header is None:
            continue
        n = sum(1 for row in grid[header + 1:] if isinstance(_cell(row, 2), datetime.datetime))
        if n:
            out.append({"name": name, "rows": n})
    return out


def pick_sheet(path, prefer=None):
    """Choose the ledger sheet whose name shares the most words with `prefer`
    (the counterparty we are reconciling). Ties and no-match fall back to the
    first ledger sheet, which is the single-sheet case."""
    found = sheets(path)
    if not found:
        return None, []
    if prefer:
        want = set(_words(prefer))
        scored = sorted(found, key=lambda s: (-len(want & set(_words(s["name"]))), -s["rows"]))
        if want & set(_words(scored[0]["name"])):
            return scored[0]["name"], found
    return found[0]["name"], found


def _words(text):
    """Significant words in a company name — the legal-form suffixes are noise."""
    noise = {"PVT", "PVT.", "PRIVATE", "LTD", "LTD.", "LIMITED", "INC", "INC.", "LT", "CO", "AND", "&"}
    return [w for w in str(text).upper().replace(".", " ").split() if w not in noise]


def _grid_xlsx(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _grid_xls(path, sheet_name):
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
    grid = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    # normalise dates here so the row logic below is format-agnostic
    for row in grid:
        if len(row) > 2 and isinstance(row[2], (int, float)) and row[2] > 0:
            try:
                row[2] = xlrd.xldate_as_datetime(row[2], wb.datemode)
            except (ValueError, OverflowError):
                pass
    return grid


def _cell(row, i):
    return row[i] if i < len(row) else None


def parse(path, sheet_name=None, prefer=None):
    if sheet_name is None:
        sheet_name, _ = pick_sheet(path, prefer)
    grid = _grid_xlsx(path, sheet_name) if _is_xlsx(path) else _grid_xls(path, sheet_name)
    if not grid:
        raise ValueError(f"{path} has no rows")

    letterhead = str(_cell(grid[0], 0) or "").strip()

    header_row = _header_row(grid)
    if header_row is None:
        raise ValueError(f"could not find 'Voucher No.' header row in {path}")

    rows = []
    for row in grid[header_row + 1:]:
        date = _cell(row, 2)
        if not isinstance(date, datetime.datetime):
            continue  # opening-balance label row, blank rows, etc.
        dr, cr = _cell(row, 6), _cell(row, 7)
        rows.append({
            "date": date.strftime("%Y-%m-%d"),
            "voucher_no": _clean(_cell(row, 0)),
            "vch_type": _clean(_cell(row, 1)),
            "particulars": _clean(_cell(row, 3)),
            "bill_no": _clean(_cell(row, 4)),
            "amount": dr if dr not in (None, "") else (cr if cr not in (None, "") else None),
            "side": "DR" if dr not in (None, "") else "CR",
            "sheet": sheet_name,
        })
    return letterhead, rows


def _clean(v):
    if v in (None, ""):
        return None
    return str(v).strip()
