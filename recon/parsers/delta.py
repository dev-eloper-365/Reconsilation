"""Parser for Delta Global Resources' Tally ledger export (.xlsx).
See profiles/delta.md for the column layout and quirks this encodes.
"""
import datetime
import openpyxl


def letterhead(path, sheet_name=None):
    """Row 1 col A — the company whose books these are. Check it before
    trusting a file: a mislabelled export matches nothing and looks like a
    reconciliation finding rather than the file mix-up it is."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    return str(ws.cell(row=1, column=1).value or "").strip()


def parse(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Date":
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"could not find 'Date' header row in {path}")

    # The columnar register export also has a "Date" header in column A but a
    # completely different column layout — reading it as a ledger yields silent
    # nonsense, so refuse it by name instead.
    if str(ws.cell(row=header_row, column=6).value or "").strip() != "Debit":
        raise ValueError(
            f"{path} does not look like the Ledger Account export: column 6 of the header "
            f"row is '{ws.cell(row=header_row, column=6).value}', expected 'Debit'. "
            f"If this is the columnar register, it belongs in the register slot."
        )

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        date = ws.cell(row=r, column=1).value
        if not isinstance(date, datetime.datetime):
            continue  # closing balance / totals rows
        debit = ws.cell(row=r, column=6).value
        credit = ws.cell(row=r, column=7).value
        rows.append({
            "date": date.strftime("%Y-%m-%d"),
            "drcr": ws.cell(row=r, column=2).value,
            "particulars": ws.cell(row=r, column=3).value,
            "vch_type": ws.cell(row=r, column=4).value,
            "bill_no": _clean(ws.cell(row=r, column=5).value),
            "amount": debit if debit is not None else credit,
        })
    return rows


def _clean(v):
    if v is None:
        return None
    return str(v).strip()
