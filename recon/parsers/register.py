"""Parser for Delta's Tally columnar register export (.xlsx).

Same vouchers as the plain ledger export, but with the columns the ledger
lacks: quantity, rate, and the taxable Value — plus one column per ledger the
voucher touched (GST, compensation cess, TDS receivable, shortage, ...).

Columns are resolved by header text, never by position: which ledger columns
exist depends on what was posted in the period, so a different period's export
has a different column count. See profiles/register.md.
"""
import datetime
import re

import openpyxl

# Fixed columns, by exact header text. Only REQUIRED ones must be present —
# which optional columns a register carries depends on how it was exported, and
# refusing a file over a descriptive column it never needed is just rudeness.
REQUIRED = {
    "date": "Date",
    "vch_type": "Voucher Type",
    "voucher_no": "Voucher No.",
    "value": "Value",
    "gross_total": "Gross Total",
}
OPTIONAL = {
    "particulars": "Particulars",
    "buyer": "Buyer/Supplier",
    "gstin": "GSTIN/UIN",
    "narration": "Narration",
    "quantity": "Quantity",
    "rate": "Rate",
}
FIXED = {**REQUIRED, **OPTIONAL}

# Bookkeeping columns that are not ledgers: descriptive, or added by whoever
# merged several years' exports into one sheet.
IGNORED = {"Alt. Units", "Financial Year", "Source File",
           "Voucher Ref. No.", "Voucher Ref. Date"}

# Ledger columns worth naming, matched loosely because their headers carry
# rates and financial years that change between periods. EVERY matching column
# is summed per row, never just the first: a multi-year export has one TDS
# column per financial year ("194Q TDS RECEIVABLE 2023-24",
# "Tds Receivable 194Q F.Y. 2024-25", ...) and one GST column per rate slab
# (CGST @ 2.5%, CGST @ 9%). Taking the first would silently drop the rest.
PATTERNS = {
    "cess": r"compensation\s+cess",
    "tds": r"194\s*q|tds",
    "shortage": r"shortage",
    "rounding": r"rounding",
    "gst": r"^(cgst|sgst|igst|utgst)\b",
}


def letterhead(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    return str(ws.cell(row=1, column=1).value or "").strip()


def cess_rate_per_unit(header):
    """The per-tonne cess rate is written into the column header itself, e.g.
    'GST Compensation Cess @ 400 PMT'. Read it rather than hardcoding 400 —
    the rate is a matter of law and has changed before."""
    m = re.search(r"@\s*([\d.]+)\s*PMT", header or "", re.I)
    return float(m.group(1)) if m else None


def parse(path, sheet_name=None):
    """-> (letterhead, rows, meta). meta carries the header text of the
    detected special columns so callers can report what they matched."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    head = str(ws.cell(row=1, column=1).value or "").strip()

    # The header row is the one carrying the columns this parser needs. Plain
    # Tally registers start it in column A with "Date"; a merged multi-year sheet
    # prefixes "Financial Year" / "Source File", so anchor on the payload columns
    # instead of on a fixed position.
    header_row, headers = None, {}
    for r in range(1, min(ws.max_row, 60) + 1):
        labels = {}
        for c in range(1, ws.max_column + 1):
            label = ws.cell(row=r, column=c).value
            if label not in (None, ""):
                labels[c] = str(label).strip()
        names = set(labels.values())
        if {"Value", "Gross Total"} <= names or ("Date" in names and "Voucher No." in names):
            header_row, headers = r, labels
            break
    if header_row is None:
        raise ValueError(f"could not find a register header row in {path}")

    by_label = {v: k for k, v in headers.items()}
    cols = {name: by_label.get(label) for name, label in FIXED.items()}
    missing = [REQUIRED[n] for n in REQUIRED if cols[n] is None]
    if missing:
        raise ValueError(
            f"register is missing required column(s): {', '.join(missing)}. "
            f"It needs at least Date, Voucher Type, Voucher No., Value and Gross Total — "
            f"re-export it with the Value (taxable amount) column shown."
        )

    # every column that isn't one of the fixed ones is a ledger the voucher hit
    used = {c for c in cols.values() if c}
    ledger_cols = {c: h for c, h in headers.items() if c not in used and h not in IGNORED}
    special = {}
    for name, pattern in PATTERNS.items():
        matched = [c for c, h in ledger_cols.items() if re.search(pattern, h, re.I)]
        if matched:
            special[name] = matched

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        date = ws.cell(row=r, column=cols["date"]).value
        if not isinstance(date, datetime.datetime):
            continue  # 'Grand Total' and any other trailing summary row
        row = {name: (ws.cell(row=r, column=c).value if c else None) for name, c in cols.items()}
        row["date"] = date.strftime("%Y-%m-%d")
        for key in ("particulars", "buyer", "vch_type", "voucher_no", "gstin", "narration"):
            row[key] = str(row[key]).strip() if row[key] not in (None, "") else None
        for name, cs in special.items():
            row[name] = sum(ws.cell(row=r, column=c).value or 0 for c in cs) or None
        row["ledgers"] = {
            h: ws.cell(row=r, column=c).value
            for c, h in ledger_cols.items()
            if ws.cell(row=r, column=c).value not in (None, "")
        }
        rows.append(row)

    cess_headers = [headers[c] for c in special.get("cess", [])]
    meta = {
        "missing_optional_columns": [OPTIONAL[n] for n in OPTIONAL if cols[n] is None],
        # one entry per named group, listing every column folded into it
        "special_columns": {name: ", ".join(headers[c] for c in cs)
                            for name, cs in special.items()},
        "cess_per_unit": cess_rate_per_unit(cess_headers[0]) if cess_headers else None,
        "gst_columns": [headers[c] for c in special.get("gst", [])],
        "ledger_columns": sorted(ledger_cols.values()),
    }
    return head, rows, meta
